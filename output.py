from openpyxl  import load_workbook
wb = load_workbook('./data/实验数据.xlsx')
ws1 = wb["Sheet1"]
print(wb.sheetnames)
ws1["A2"].value = "改改试试"
cellB2_value = ws1['A2'].value
print("单元格B2内容为：", cellB2_value)
wb.save("./data/实验数据.xlsx")